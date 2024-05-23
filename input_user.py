import openpyxl
import Sheet_Create
from openpyxl.styles import PatternFill 


class InputUser:
    def __init__(self, workbook):
        self.workbook = workbook



    @staticmethod
    def sheet():
        workbook = openpyxl.Workbook()
        workbook.save('students.xlsx')



    def add_student(self):
        sheet_name = input(f'Please enter the field name {self.workbook.sheetnames}: ').lower()
        sheet = self.workbook[sheet_name]

        name = input('Enter your name: ')
        family_name = input('Enter your family name: ')
        National_Code = input('Enter your national code: ')
        Student_Number = input('Enter your student number: ')
        Field = input('Enter your field: ')
        GPA = input('Enter your GPA: ')

        sheet.append([name, family_name, National_Code, Student_Number, Field, GPA])

        self.workbook.save('students.xlsx')



    def delete_student(self):
        sheet_name = input('Please enter the field name [Computer, Chemistry, Math, Electronic]: ').lower()
        sheet = self.workbook[sheet_name]

        student_id = input('Enter the student ID to delete: ')

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=4, max_col=4):
            for cell in row:
                if cell.value == student_id:
                    sheet.delete_rows(cell.row)
                    print(f'Student with ID {student_id} deleted successfully.')
                    break
            else:
                continue
            break

        self.workbook.save('students.xlsx')



    def create_sheets(self):
        answer = input('Do you want to add a new sheet? (yes/no): ')

        if answer.lower() == 'yes':
            new_sheet_name = input('Enter the name for the new sheet: ')
            sheet = self.workbook.create_sheet(title=new_sheet_name)

            sheet['A1'] = 'Name'
            sheet['B1'] = 'Family'
            sheet['C1'] = 'National Code'
            sheet['D1'] = 'Student Number'
            sheet['E1'] = 'Field'
            sheet['F1'] = 'GPA'

            # Set column width for all columns
            for col in range(1, 7):
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

                # Apply color to row A1:F1
            for col in range(1, 7):
                cell = sheet.cell(row=1, column=col)
                cell.fill = PatternFill(start_color='FFC0CB', fill_type='solid')

            print(f'your sheet by name : {new_sheet_name} created')
            self.workbook.save('students.xlsx')
            



    def remove_sheets(self):
        answer = input('Do you want to remove sheet? (yes/no): ')

        if answer.lower() == 'yes':
            print('all sheets : ')

            for sheet in self.workbook.sheetnames:
                print(f'-{sheet}')
            
        sheet_remove = input("Enter the name of the sheet you want to remove: ")

        if sheet_remove in self.workbook.sheetnames :
            self.workbook.remove(self.workbook[sheet_remove])

            print(f'sheet {sheet_remove} removed successfully.')
            self.workbook.save('students.xlsx')
        
        else:
            print(f"Sheet '{sheet_remove}' not found.")

    


    def best_of_student(self):
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            
            # مرتب سازی دانشجویان بر اساس معدل
            data = []
            for row in range(2, sheet.max_row+1):
                name = sheet.cell(row=row, column=1).value
                family_name = sheet.cell(row=row, column=2).value
                GPA = float(sheet.cell(row=row, column=6).value)
                data.append((name, family_name, GPA))
            
            data.sort(key=lambda x: x[2], reverse=True)
            
            # نمایش بهترین دانشجو
            print(f"Best student in {sheet_name} field:")
            print(f"Name: {data[0][0]}")
            print(f"Family: {data[0][1]}")
            print(f"GPA: {data[0][2]}")
            print()
    

    
    
    def edit_student(self):
        sheet_name = input('Please enter the field name [Computer, Chemistry, Math, Electronic]: ').lower()
        sheet = self.workbook[sheet_name]

        student_id = input('Enter the student ID to edit: ')

        for row in range(2, sheet.max_row+1):
            student_number = sheet.cell(row=row, column=4).value

            if student_number == student_id:
                name = input(f'Enter new name (current: {sheet.cell(row=row, column=1).value}): ')
                family_name = input(f'Enter new family name (current: {sheet.cell(row=row, column=2).value}): ')
                national_code = input(f'Enter new national code (current: {sheet.cell(row=row, column=3).value}): ')
                field = input(f'Enter new field (current: {sheet.cell(row=row, column=5).value}): ')
                gpa = input(f'Enter new GPA (current: {sheet.cell(row=row, column=6).value}): ')

                sheet.cell(row=row, column=1).value = name
                sheet.cell(row=row, column=2).value = family_name
                sheet.cell(row=row, column=3).value = national_code
                sheet.cell(row=row, column=5).value = field
                sheet.cell(row=row, column=6).value = gpa

                print(f'Student with ID {student_id} edited successfully.')
                self.workbook.save('students.xlsx')
                return

        print(f'Student with ID {student_id} not found.')



        


# create_sheet = Sheet_Create.CreateSheet()
# create_sheet.sheet()

workbook = openpyxl.load_workbook('students.xlsx')
create_user = InputUser(workbook)
# create_user.add_student()
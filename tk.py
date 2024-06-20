import tkinter as tk
from tkinter import ttk, simpledialog, messagebox
import openpyxl
from openpyxl.styles import PatternFill
from PIL import Image, ImageTk
from tkinter import Tk, Canvas, Frame, Label, Button, ttk
from PIL import Image, ImageTk
for i


class AnimatedGIF(Label):
    def __init__(self, master, path, speed):
        self.path = path
        self.speed = speed
        self.frames = []

        img = Image.open(self.path)
        for frame in range(0, img.n_frames):
            img.seek(frame)
            frame_img = ImageTk.PhotoImage(img.copy())
            self.frames.append(frame_img)

        super().__init__(master)
        self.delay = speed
        self.idx = 0
        self.cancel = self.after(self.delay, self.update_frames)
        self.pack()

    def update_frames(self):
        self.config(image=self.frames[self.idx])
        self.idx += 1
        if self.idx == len(self.frames):
            self.idx = 0
        self.cancel = self.after(self.delay, self.update_frames)

class StudentManagementGUI:
    def __init__(self, workbook):
        self.workbook = workbook
        self.root = Tk()
        self.root.title("Student Management System")

        # Make the window full screen and remove title bar
        self.root.attributes('-fullscreen', True)
        self.root.overrideredirect(True)

        # Create a canvas to hold the background image
        self.canvas = Canvas(self.root, width=1920, height=1080)
        self.canvas.pack(fill='both', expand=True)

        # Load the background GIF
        self.background_gif = AnimatedGIF(self.canvas, "bpxxqqvps4h91.gif", speed=100)
        self.background_gif.pack(fill='both', expand=True)

        # Create a frame to hold the menu buttons
        self.menu_frame = Frame(self.canvas, bg='blue')
        self.menu_frame.place(relx=0.5, rely=0.05, anchor='n')  # Change rely to place buttons higher

        # Create a style object
        self.style = ttk.Style()

        # Configure the style for the menu buttons
        self.style.configure('MenuButton.TButton', background='yellow', foreground='black', font=('Arial', 12, 'bold'))

        self.menu_buttons = []
        for i in range(1, 7):
            button = ttk.Button(self.menu_frame, text=self.get_menu_option(i), command=lambda x=i: self.handle_menu_choice(x), style='MenuButton.TButton')
            button.grid(row=0, column=i-1, padx=10, pady=10)
            self.menu_buttons.append(button)

        # Exit button
        exit_button = Button(self.canvas, text="Exit", command=self.root.quit, bg='yellow', fg='black', font=('Arial', 12, 'bold'))
        exit_button.place(relx=0.95, rely=0.05, anchor='ne')

        self.operation_frame = None
        self.operation_widgets = {}

    def get_menu_option(self, choice):
        options = {
            1: "Add Student",
            2: "Delete Student",
            3: "Create Sheets",
            4: "Remove Sheets",
            5: "Best Student",
            6: "Edit Student"
        }
        return options.get(choice, "")
    

    def handle_menu_choice(self, choice):
        if self.operation_frame:
            self.operation_frame.destroy()

        if choice == 1:
            self.create_add_student_widgets()
        elif choice == 2:
            self.create_delete_student_widgets()
        elif choice == 3:
            self.create_create_sheets_widgets()
        elif choice == 4:
            self.create_remove_sheets_widgets()
        elif choice == 5:
            self.create_best_of_student_widgets()
        elif choice == 6:
            self.create_edit_student_widgets()

    def create_add_student_widgets(self):
        # Create a new Toplevel window
        add_student_window = tk.Toplevel(self.root)
        add_student_window.title("Add Student")

        # Make the window full screen and remove title bar
        add_student_window.attributes('-fullscreen', True)
        add_student_window.overrideredirect(True)

        # Create a canvas to hold the background GIF
        canvas = tk.Canvas(add_student_window, width=1920, height=1080)
        canvas.pack(fill=tk.BOTH, expand=True)

        # Load the background GIF
        background_gif = AnimatedGIF(canvas, "0k6meqvps4h91.gif", speed=100)
        background_gif.pack(fill=tk.BOTH, expand=True)

        # Create a frame to hold the widgets
        frame = tk.Frame(canvas, bg='blue')  # Set the background color to gray
        frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

        # Create widgets on the frame
        labels = ["Sheet Name:", "Name:", "Family Name:", "National Code:", "Student Number:", "Field:", "GPA:"]
        for i, label in enumerate(labels):
            label_widget = tk.Label(frame, text=label, bg='blue', fg='black')  # Set the label background to gray and text to black
            label_widget.grid(row=i, column=0, padx=10, pady=5)

            if label == "Sheet Name:":
                self.operation_widgets["sheet_name"] = ttk.Combobox(frame, values=self.workbook.sheetnames)
                self.operation_widgets["sheet_name"].grid(row=i, column=1, padx=10, pady=5)
            else:
                entry_widget = ttk.Entry(frame)
                entry_widget.grid(row=i, column=1, padx=10, pady=5)
                self.operation_widgets[label.replace(":", "").lower()] = entry_widget

        add_button = tk.Button(frame, text="Add Student", command=self.add_student, bg='yellow', fg='black')  # Set the button background to yellow and text to black
        add_button.grid(row=len(labels), column=0, columnspan=2, pady=10)

        # Exit button for add student window
        exit_button = tk.Button(add_student_window, text="Exit", command=add_student_window.destroy, bg='yellow', fg='black', font=('Arial', 12, 'bold'))
        exit_button.place(relx=0.95, rely=0.05, anchor=tk.CENTER)


    def add_student(self):
        sheet_name = self.operation_widgets["sheet_name"].get()
        name = self.operation_widgets["name"].get()
        family_name = self.operation_widgets["family name"].get()
        national_code = self.operation_widgets["national code"].get()
        student_number = self.operation_widgets["student number"].get()
        field = self.operation_widgets["field"].get()
        gpa = self.operation_widgets["gpa"].get()

        if not all([sheet_name, name, family_name, national_code, student_number, field, gpa]):
            messagebox.showerror("Input Error", "All fields are required.")
            return

        try:
            gpa = float(gpa)
        except ValueError:
            messagebox.showerror("Input Error", "GPA must be a number.")
            return

        sheet = self.workbook[sheet_name]
        new_row = [name, family_name, national_code, student_number, field, gpa]
        sheet.append(new_row)
        self.workbook.save("students.xlsx")

        messagebox.showinfo("Success", "Student added successfully.")

    def create_delete_student_widgets(self):
        # Create a new Toplevel window
        delete_student_window = tk.Toplevel(self.root)
        delete_student_window.title("Delete Student")

        # Make the window full screen and remove title bar
        delete_student_window.attributes('-fullscreen', True)
        delete_student_window.overrideredirect(True)

        # Create a canvas to hold the background GIF
        canvas = tk.Canvas(delete_student_window, width=1920, height=1080)
        canvas.pack(fill=tk.BOTH, expand=True)

        # Load the background GIF
        background_gif = AnimatedGIF(canvas, "lhnvupvps4h91.gif", speed=100)
        background_gif.pack(fill=tk.BOTH, expand=True)

        # Create a frame to hold the widgets
        frame = tk.Frame(canvas, bg='Gray')  # Set the background color to gray
        frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

        # Create widgets on the frame
        labels = ["Sheet Name:", "Student Number:"]
        for i, label in enumerate(labels):
            label_widget = tk.Label(frame, text=label, bg='Gray', fg='black')  # Set the label background to gray and text to black
            label_widget.grid(row=i, column=0, padx=10, pady=5)

            if label == "Sheet Name:":
                self.operation_widgets["sheet_name"] = ttk.Combobox(frame, values=self.workbook.sheetnames)
                self.operation_widgets["sheet_name"].grid(row=i, column=1, padx=10, pady=5)
            else:
                entry_widget = ttk.Entry(frame)
                entry_widget.grid(row=i, column=1, padx=10, pady=5)
                if label == "Student Number:":
                    self.operation_widgets["student_number"] = entry_widget
                else:
                    self.operation_widgets[label.replace(":", "").lower()] = entry_widget

        delete_button = tk.Button(frame, text="Delete Student", command=self.delete_student, bg='yellow', fg='black')  # Set the button background to yellow and text to black
        delete_button.grid(row=len(labels), column=0, columnspan=2, pady=10)

        # Exit button for delete student window
        exit_button = tk.Button(delete_student_window, text="Exit", command=delete_student_window.destroy, bg='yellow', fg='black', font=('Arial', 12, 'bold'))
        exit_button.place(relx=0.95, rely=0.05, anchor=tk.CENTER)

    def delete_student(self):
        sheet_name = self.operation_widgets["sheet_name"].get()
        sheet = self.workbook[sheet_name]

        student_number = self.operation_widgets["student_number"].get()  # Changed from "student_id"

        for row in range(2, sheet.max_row + 1):
            student_number_cell = sheet.cell(row=row, column=4).value  # Assuming student number is in column D (4th column)
            if student_number_cell == student_number:
                sheet.delete_rows(row)
                print(f'Student with ID {student_number} deleted successfully.')
                break

        self.workbook.save('students.xlsx')

        self.operation_widgets["student_number"].delete(0, tk.END)
        tk.messagebox.showinfo("Student Deleted", "Student deleted successfully.")



        

    def create_create_sheets_widgets(self):
        # Create a new Toplevel window
        create_sheets_window = tk.Toplevel(self.root)
        create_sheets_window.title("Create Sheets")

        # Make the window full screen and remove title bar
        create_sheets_window.attributes('-fullscreen', True)
        create_sheets_window.overrideredirect(True)

        # Create a canvas to hold the background GIF
        canvas = tk.Canvas(create_sheets_window, width=1920, height=1080)
        canvas.pack(fill=tk.BOTH, expand=True)

        # Load the background GIF
        background_gif = AnimatedGIF(canvas, "helia-goyer-brb.gif", speed=100)
        background_gif.pack(fill=tk.BOTH, expand=True)

        # Create a frame to hold the widgets
        frame = tk.Frame(canvas, bg='blue')  # Set the background color to gray
        frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

        # Create widgets on the frame
        sheet_name_label = tk.Label(frame, text="Sheet Name:", bg='blue', fg='black')  # Set the label background to gray and text to black
        sheet_name_label.grid(row=0, column=0, padx=10, pady=5)

        self.operation_widgets["sheet_name"] = ttk.Entry(frame)
        self.operation_widgets["sheet_name"].grid(row=0, column=1, padx=10, pady=5)

        create_button = tk.Button(frame, text="Create Sheet", command=self.create_sheets, bg='yellow', fg='black')  # Set the button background to yellow and text to black
        create_button.grid(row=1, column=0, columnspan=2, pady=10)

        # Exit button for create sheets window
        exit_button = tk.Button(create_sheets_window, text="Exit", command=create_sheets_window.destroy, bg='yellow', fg='black', font=('Arial', 12, 'bold'))
        exit_button.place(relx=0.95, rely=0.05, anchor=tk.CENTER)

    def create_sheets(self):
        new_sheet_name = self.operation_widgets["sheet_name"].get()
        sheet = self.workbook.create_sheet(title=new_sheet_name)

        sheet['A1'] = 'Name'
        sheet['B1'] = 'Family'
        sheet['C1'] = 'National Code'
        sheet['D1'] = 'Student Number'
        sheet['E1'] = 'Field'
        sheet['F1'] = 'GPA'

        for col in range(1, 7):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

        for col in range(1, 7):
            cell = sheet.cell(row=1, column=col)
            cell.fill = PatternFill(start_color='FFC0CB', fill_type='solid')

        self.workbook.save('students.xlsx')

        self.operation_widgets["sheet_name"].delete(0, tk.END)
        tk.messagebox.showinfo("Sheet Created", "New sheet created successfully.")





    def create_remove_sheets_widgets(self):
        # Create a new Toplevel window
        remove_sheets_window = tk.Toplevel(self.root)
        remove_sheets_window.title("Remove Sheets")

        # Make the window full screen and remove title bar
        remove_sheets_window.attributes('-fullscreen', True)
        remove_sheets_window.overrideredirect(True)

        # Create a canvas to hold the background GIF
        canvas = tk.Canvas(remove_sheets_window, width=1920, height=1080)
        canvas.pack(fill=tk.BOTH, expand=True)

        # Load the background GIF
        background_gif = AnimatedGIF(canvas, "tumblr_13d2c753eed929097cc13bbb1d3e482c_fe67f6e7_1920.gif", speed=100)  # Add your GIF path here
        background_gif.pack(fill=tk.BOTH, expand=True)

        # Create a frame to hold the widgets
        frame = tk.Frame(canvas, bg='blue')  # Set the background color to gray
        frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

        # Create widgets on the frame
        sheet_name_label = tk.Label(frame, text="Sheet Name:", bg='blue', fg='black')  # Set the label background to gray and text to black
        sheet_name_label.grid(row=0, column=0, padx=10, pady=5)

        self.operation_widgets["sheet_name"] = ttk.Combobox(frame, values=self.workbook.sheetnames)
        self.operation_widgets["sheet_name"].grid(row=0, column=1, padx=10, pady=5)

        remove_button = tk.Button(frame, text="Remove Sheet", command=self.remove_sheets, bg='yellow', fg='black')  # Set the button background to yellow and text to black
        remove_button.grid(row=1, column=0, columnspan=2, pady=10)

        # Exit button for remove sheets window
        exit_button = tk.Button(remove_sheets_window, text="Exit", command=remove_sheets_window.destroy, bg='yellow', fg='black', font=('Arial', 12, 'bold'))
        exit_button.place(relx=0.95, rely=0.05, anchor=tk.CENTER)

    def remove_sheets(self):
        sheet_name = self.operation_widgets["sheet_name"].get()

        if sheet_name in self.workbook.sheetnames:
            self.workbook.remove(self.workbook[sheet_name])
            self.workbook.save('students.xlsx')
            tk.messagebox.showinfo("Sheet Removed", f"Sheet '{sheet_name}' removed successfully.")
        else:
            tk.messagebox.showerror("Error", f"Sheet '{sheet_name}' not found.")





    def create_best_of_student_widgets(self):
        # Destroy any existing operation frame
        if self.operation_frame:
            self.operation_frame.destroy()

        # Create a new Toplevel window
        best_student_window = tk.Toplevel(self.root)
        best_student_window.title("Best Students")

        # Make the window full screen and remove title bar
        best_student_window.attributes('-fullscreen', True)
        best_student_window.overrideredirect(True)

        # Create a canvas to hold the background GIF
        canvas = tk.Canvas(best_student_window, width=1920, height=1080)
        canvas.pack(fill=tk.BOTH, expand=True)

        # Load the background GIF
        background_gif = AnimatedGIF(canvas, "final+with+character+in+window.gif", speed=100)  # Add your GIF path here
        background_gif.pack(fill=tk.BOTH, expand=True)

        # Create a frame to hold the widgets
        frame = tk.Frame(canvas, bg='lightblue')  # Set the background color to light blue
        frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

        # Create a Label for each field
        fields_label = tk.Label(frame, text="Best Students in Each Field", bg='lightblue', fg='navy', font=('Arial', 18, 'bold'))
        fields_label.grid(row=0, column=0, padx=10, pady=10, columnspan=2)

        row_num = 1
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            data = []
            for row in range(2, sheet.max_row + 1):
                name = sheet.cell(row=row, column=1).value
                family_name = sheet.cell(row=row, column=2).value
                GPA = float(sheet.cell(row=row, column=6).value)
                data.append((name, family_name, GPA))

            data.sort(key=lambda x: x[2], reverse=True)

            # Create a Label for each field's best student
            field_label = tk.Label(frame, text=f"Best student in {sheet_name} field:", bg='lightblue', fg='navy', font=('Arial', 14, 'bold'))
            field_label.grid(row=row_num, column=0, padx=10, pady=10, sticky='w')

            student_info_label = tk.Label(frame, text=f"Name: {data[0][0]}, Family: {data[0][1]}, GPA: {data[0][2]}", bg='lightblue', fg='black', font=('Arial', 12))
            student_info_label.grid(row=row_num, column=1, padx=10, pady=10, sticky='w')

            row_num += 1

        # Exit button for best student window
        exit_button = tk.Button(frame, text="Exit", command=best_student_window.destroy, bg='yellow', fg='black', font=('Arial', 12, 'bold'))
        exit_button.grid(row=row_num, column=0, columnspan=2, pady=10)



        

    def create_edit_student_widgets(self):
        # Create a new Toplevel window
        edit_student_window = tk.Toplevel(self.root)
        edit_student_window.title("Edit Student")
        edit_student_window.attributes('-fullscreen', True)
        edit_student_window.overrideredirect(True)

        # Create a canvas to hold the background GIF
        canvas = tk.Canvas(edit_student_window, width=1920, height=1080)
        canvas.pack(fill=tk.BOTH, expand=True)

        # Load the background GIF
        background_gif = AnimatedGIF(canvas, "137e45c33cb5939abcd3ff4d8f858c63.gif", speed=100)  # Add your GIF path here
        background_gif.pack(fill=tk.BOTH, expand=True)

        # Create a frame to hold the widgets
        frame = tk.Frame(canvas, bg='lightgreen')  # Set the background color to light green
        frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

        # Create widgets on the frame
        labels = ["Sheet Name:", "Student ID:", "Name:", "Family Name:", "National Code:", "Field:", "GPA:"]
        for i, label in enumerate(labels):
            label_widget = tk.Label(frame, text=label, bg='lightgreen', fg='black')  # Set the label background to light green and text to black
            label_widget.grid(row=i, column=0, padx=10, pady=5)

            if label == "Sheet Name:":
                self.operation_widgets["sheet_name"] = ttk.Combobox(frame, values=self.workbook.sheetnames)
                self.operation_widgets["sheet_name"].grid(row=i, column=1, padx=10, pady=5)
            else:
                entry_widget = ttk.Entry(frame)
                entry_widget.grid(row=i, column=1, padx=10, pady=5)
                self.operation_widgets[label.replace(":", "").lower()] = entry_widget

        edit_button = tk.Button(frame, text="Edit Student", command=self.edit_student, bg='yellow', fg='black')  # Set the button background to yellow and text to black
        edit_button.grid(row=len(labels), column=0, columnspan=2, pady=10)

        # Exit button for edit student window
        exit_button = tk.Button(edit_student_window, text="Exit", command=edit_student_window.destroy, bg='yellow', fg='black', font=('Arial', 12, 'bold'))
        exit_button.place(relx=0.95, rely=0.05, anchor=tk.CENTER)

    def edit_student(self):
        sheet_name = self.operation_widgets["sheet_name"].get()
        sheet = self.workbook[sheet_name]

        student_id = self.operation_widgets["student_id"].get()

        for row in range(2, sheet.max_row + 1):
            student_number = sheet.cell(row=row, column=4).value

            if student_number == student_id:
                name = self.operation_widgets["name"].get()
                family_name = self.operation_widgets["family_name"].get()
                national_code = self.operation_widgets["national_code"].get()
                field = self.operation_widgets["field"].get()
                gpa = self.operation_widgets["gpa"].get()

                sheet.cell(row=row, column=1).value = name
                sheet.cell(row=row, column=2).value = family_name
                sheet.cell(row=row, column=3).value = national_code
                sheet.cell(row=row, column=5).value = field
                sheet.cell(row=row, column=6).value = gpa

                self.workbook.save('students.xlsx')

                self.operation_widgets["sheet_name"].set('')
                self.operation_widgets["student_id"].delete(0, tk.END)
                self.operation_widgets["name"].delete(0, tk.END)
                self.operation_widgets["family_name"].delete(0, tk.END)
                self.operation_widgets["national_code"].delete(0, tk.END)
                self.operation_widgets["field"].delete(0, tk.END)
                self.operation_widgets["gpa"].delete(0, tk.END)

                tk.messagebox.showinfo("Student Edited", f"Student with ID {student_id} edited successfully.")
                return

        tk.messagebox.showerror("Error", f"Student with ID {student_id} not found.")

    def run(self):
        self.root.mainloop()

workbook = openpyxl.load_workbook('students.xlsx')
gui = StudentManagementGUI(workbook)
gui.run()
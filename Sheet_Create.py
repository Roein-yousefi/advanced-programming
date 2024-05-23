import openpyxl
from openpyxl.styles import PatternFill 

wb = openpyxl.load_workbook('students.xlsx')

class CreateSheet:
    def sheet(self):
        sheet_created = False
        
        for sheet_name in wb.sheetnames:
            if sheet_name not in wb.sheetnames:
                sheet = wb.create_sheet(sheet_name)

                sheet['A1'] = 'Name'
                sheet['B1'] = 'Family'
                sheet['C1'] = 'National Code'
                sheet['D1'] = 'Student Number'
                sheet['E1'] = 'Field'
                sheet['F1'] = 'GPA'

                # Set column width for all columns
                for col in range(1, 7):
                    sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

                # Apply color to row A1:F1
                for col in range(1, 7):
                    cell = sheet.cell(row=1, column=col)
                    cell.fill = PatternFill(start_color='FFC0CB', fill_type='solid')

                sheet_created = True
        
        if sheet_created:
            wb.save('students.xlsx')

# CreateSheet = CreateSheet()
# CreateSheet.sheet()
